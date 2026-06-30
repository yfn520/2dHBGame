"""创建技能配置 Excel 文件"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "excel", "skills.xlsx")

HEADERS = [
    ("id", "ID", 8),
    ("name", "名称", 12),
    ("type", "类型", 12),
    ("damage_ratio", "伤害倍率", 10),
    ("cooldown", "冷却时间", 10),
    ("animation", "动画名", 12),
    ("range", "范围/射程", 10),
    ("projectile_scene", "弹道场景", 24),
    ("max_pierce", "穿透次数", 10),
    ("aoe_radius", "AOE半径", 10),
    ("buff_on_hit", "命中BuffID", 10),
    ("buff_chance", "Buff概率", 10),
    ("buff_on_self", "自身BuffID", 10),
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 示例技能数据
ROWS = [
    [1001, "普攻", "melee", 1.0, 0.5, "attack", 40, "", 0, 0, 0, 0, 0],
    [1002, "火球术", "projectile", 1.5, 3.0, "skill1", 300, "res://scenes/effects/fireball.tscn", 0, 0, 2001, 1.0, 0],
    [1003, "旋风斩", "aoe", 2.0, 5.0, "skill2", 0, "", 0, 80, 0, 0, 1005],
    [1004, "冰霜箭", "penetrate", 1.2, 4.0, "skill3", 400, "res://scenes/effects/frost_arrow.tscn", -1, 0, 1003, 0.5, 0],
    [1005, "全屏斩", "fullscreen", 3.0, 10.0, "skill2", 0, "", 0, 0, 0, 0, 0],
]


def create():
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"
    for col, (_, title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    for row_idx, row_data in enumerate(ROWS, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_PATH)
    print(f"已创建: {EXCEL_PATH}")


if __name__ == "__main__":
    create()
