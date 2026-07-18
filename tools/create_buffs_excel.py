"""创建Buff配置 Excel 文件"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "excel", "buffs.xlsx")

HEADERS = [
    ("id", "ID", 8),
    ("name", "名称", 12),
    ("type", "类型", 12),
    ("duration", "持续时间", 10),
    ("interval", "跳数间隔", 10),
    ("tick_damage", "每跳伤害", 10),
    ("slow_ratio", "减速比例", 10),
    ("max_stacks", "最大层数", 10),
    ("effect_scene", "特效场景", 28),
]

HEADER_FILL = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ROWS = [
    [1001, "中毒", "poison", 5.0, 1.0, 5, 0, 5, "res://assets/effects/poison_fx.tscn"],
    [1002, "燃烧", "burn", 3.0, 0.5, 8, 0, 1, "res://assets/effects/burn_fx.tscn"],
    [1003, "冰冻", "freeze", 2.0, 0, 0, 0, 1, "res://assets/effects/freeze_fx.tscn"],
    [1004, "麻痹", "paralysis", 1.5, 0, 0, 0, 1, "res://assets/effects/paralysis_fx.tscn"],
    [1005, "无敌", "invincible", 3.0, 0, 0, 0, 1, "res://assets/effects/invincible_fx.tscn"],
    [1006, "眩晕", "stun", 2.0, 0, 0, 0, 1, "res://assets/effects/stun_fx.tscn"],
    [1007, "减速", "slow", 4.0, 0, 0, 0.5, 1, ""],
]


def create():
    wb = Workbook()
    ws = wb.active
    ws.title = "Buffs"
    for col, (_, title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    for row_idx, row_data in enumerate(ROWS, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_PATH)
    print(f"已创建: {EXCEL_PATH}")


if __name__ == "__main__":
    create()
