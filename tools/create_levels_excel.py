"""创建关卡配置 Excel 文件"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "excel", "levels.xlsx")

HEADERS = [
    ("id", "ID", 10),
    ("name", "名称", 14),
    ("scene_path", "场景路径", 36),
    ("spawn_x", "出生点X", 10),
    ("spawn_y", "出生点Y", 10),
    ("bgm", "背景音乐", 24),
    ("description", "描述", 30),
]

HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ROWS = [
    [1, "丛林", "res://scenes/jungle_01.tscn", 160, 350, "", "初始关卡"],
    [2, "星地", "res://scenes/xing.tscn", 160, 350, "", "第二关卡"],
]


def create():
    wb = Workbook()
    ws = wb.active
    ws.title = "Levels"
    for col, (_, title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width
    for row_idx, row_data in enumerate(ROWS, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_PATH)
    print(f"已创建: {EXCEL_PATH}")


if __name__ == "__main__":
    create()
