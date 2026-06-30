"""创建物品配置 Excel 文件 (data/excel/items.xlsx)

运行方式:
    python tools/create_items_excel.py

此脚本仅用于初始化 Excel 文件，后续直接用 Excel 编辑即可。
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "excel", "items.xlsx")
JSON_PATH = os.path.join(PROJECT_ROOT, "data", "items.json")

# Excel 表头定义
HEADERS = [
    ("id", "ID", 10),
    ("name", "名称", 14),
    ("type", "类型", 12),
    ("description", "描述", 30),
    ("stackable", "可堆叠", 10),
    ("max_count", "最大数量", 10),
    ("attack", "攻击力", 10),
    ("defense", "防御力", 10),
    ("max_hp", "最大生命", 10),
    ("move_speed", "移动速度", 12),
    ("heal_amount", "回复量", 10),
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_excel_from_json():
    """从现有 items.json 生成 items.xlsx"""
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    # 写表头
    for col, (_, title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    # 写数据
    for row_idx, (item_id, item) in enumerate(data.items(), 2):
        stats = item.get("stats", {})
        values = [
            int(item_id),
            item.get("name", ""),
            item.get("type", ""),
            item.get("description", ""),
            1 if item.get("stackable", False) else 0,
            item.get("max_count", 1),
            stats.get("attack", 0),
            stats.get("defense", 0),
            stats.get("max_hp", 0),
            stats.get("move_speed", 0),
            item.get("heal_amount", 0),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_PATH)
    print(f"Excel 文件已创建: {EXCEL_PATH}")


if __name__ == "__main__":
    create_excel_from_json()
