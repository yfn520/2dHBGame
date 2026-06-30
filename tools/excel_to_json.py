"""通用 Excel → JSON 转换工具

扫描 data/excel/ 目录下所有 .xlsx 文件，转换为 data/ 目录下的 .json 文件。
每个 Excel 文件的第一行为表头，第二行起为数据。

运行方式:
    python tools/excel_to_json.py              # 转换所有 Excel
    python tools/excel_to_json.py items        # 只转换 items.xlsx

约定:
    - Excel 第一行为表头(中文或英文均可)
    - 第一列必须是 id 列(值为整数)
    - 表头名称会映射为 JSON key
    - 值为 1/0 的列自动转为 true/false (仅限 stackable 等布尔字段)
    - 值为数字的列自动转为 int 或 float
    - stats 子对象: attack, defense, max_hp, move_speed 会自动归入 stats 字典

新增配置表只需:
    1. 在 data/excel/ 下创建新的 .xlsx
    2. 按上面约定填写表头和数据
    3. 运行 python tools/excel_to_json.py
"""
import os
import sys
import json
import math
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "data", "excel")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "data")

# 这些字段会归入 stats 子对象
STAT_FIELDS = {"attack", "defense", "max_hp", "move_speed"}
# 这些字段值为 1/0 时转为 true/false
BOOL_FIELDS = {"stackable"}

# 中文表头 → 英文 key 的映射
HEADER_MAP = {
    "ID": "id",
    "名称": "name",
    "类型": "type",
    "描述": "description",
    "可堆叠": "stackable",
    "最大数量": "max_count",
    "攻击力": "attack",
    "防御力": "defense",
    "最大生命": "max_hp",
    "移动速度": "move_speed",
    "回复量": "heal_amount",
    # 技能表
    "伤害倍率": "damage_ratio",
    "冷却时间": "cooldown",
    "动画名": "animation",
    "范围/射程": "range",
    "弹道场景": "projectile_scene",
    "穿透次数": "max_pierce",
    "AOE半径": "aoe_radius",
    "命中BuffID": "buff_on_hit",
    "Buff概率": "buff_chance",
    "自身BuffID": "buff_on_self",
    # Buff表
    "持续时间": "duration",
    "跳数间隔": "interval",
    "每跳伤害": "tick_damage",
    "减速比例": "slow_ratio",
    "最大层数": "max_stacks",
    "特效场景": "effect_scene",
    # 关卡表
    "场景路径": "scene_path",
    "出生点X": "spawn_x",
    "出生点Y": "spawn_y",
    "背景音乐": "bgm",
}


def parse_value(header_key: str, raw_value) -> any:
    """根据字段类型解析值"""
    if raw_value is None:
        return 0 if header_key in STAT_FIELDS else ""

    # 布尔字段
    if header_key in BOOL_FIELDS:
        try:
            return bool(int(raw_value))
        except (ValueError, TypeError):
            return False

    # 数字字段
    if isinstance(raw_value, (int, float)):
        if header_key in STAT_FIELDS or header_key in ("max_count", "heal_amount"):
            if isinstance(raw_value, float) and raw_value == int(raw_value):
                return int(raw_value)
            return raw_value
        return raw_value

    # 尝试转数字
    try:
        val = float(raw_value)
        if val == int(val):
            return int(val)
        return val
    except (ValueError, TypeError):
        return str(raw_value).strip()


def convert_excel_to_json(excel_name: str) -> bool:
    """转换单个 Excel 文件为 JSON"""
    excel_path = os.path.join(EXCEL_DIR, f"{excel_name}.xlsx")
    if not os.path.exists(excel_path):
        print(f"  跳过: {excel_path} 不存在")
        return False

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print(f"  跳过: {excel_name}.xlsx 只有表头没有数据")
        wb.close()
        return False

    # 解析表头
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    headers = []
    for h in raw_headers:
        key = HEADER_MAP.get(h, h)
        headers.append(key)

    # 解析数据行
    result = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        item_id = str(int(row[0]))
        item = {}
        stats = {}

        for col_idx, key in enumerate(headers):
            if col_idx >= len(row):
                break
            if key == "id":
                continue
            val = parse_value(key, row[col_idx])
            if key in STAT_FIELDS:
                if val and val != 0:
                    stats[key] = val
            else:
                item[key] = val

        if stats:
            item["stats"] = stats
        result[item_id] = item

    wb.close()

    # 写 JSON
    json_path = os.path.join(OUTPUT_DIR, f"{excel_name}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent="\t")
    print(f"  已生成: {json_path}")
    return True


def main():
    # 确定要转换的文件
    if len(sys.argv) > 1:
        names = sys.argv[1:]
    else:
        names = []
        for f in os.listdir(EXCEL_DIR):
            if f.endswith(".xlsx") and not f.startswith("~$"):
                names.append(f[:-5])

    if not names:
        print(f"没有找到 Excel 文件: {EXCEL_DIR}")
        return

    print(f"Excel → JSON 转换")
    print(f"  Excel 目录: {EXCEL_DIR}")
    print(f"  JSON  目录: {OUTPUT_DIR}")
    print()

    success = 0
    for name in names:
        print(f"转换: {name}.xlsx")
        if convert_excel_to_json(name):
            success += 1

    print(f"\n完成: {success}/{len(names)} 个文件已转换")


if __name__ == "__main__":
    main()
