"""通用 JSON → Excel 转换工具

扫描 data/ 目录下所有 .json 文件，转换为 data/excel/ 目录下的 .xlsx 文件。
嵌套字典会展平为列（如 effects.max_hp → effects_max_hp）。
数组用分号分隔。

运行方式:
    python tools/json_to_excel.py              # 转换所有 JSON（跳过已存在的 Excel）
    python tools/json_to_excel.py enemies      # 只转换 enemies.json
    python tools/json_to_excel.py --force      # 覆盖已存在的 Excel
"""
import os
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_DIR = os.path.join(PROJECT_ROOT, "data")
EXCEL_DIR = os.path.join(PROJECT_ROOT, "data", "excel")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 英文 key → 中文表头
KEY_TO_CN = {
    "id": "ID",
    "name": "名称",
    "type": "类型",
    "description": "描述",
	"effect_timing": "生效时机",
    "max_hp": "最大生命",
    "attack": "攻击力",
    "defense": "防御力",
    "move_speed": "移动速度",
    "stackable": "可堆叠",
    "max_count": "最大数量",
    "heal_amount": "回复量",
    "damage_ratio": "伤害倍率",
    "cooldown": "冷却时间",
    "animation": "动画名",
    "range": "范围",
    "projectile_scene": "弹道场景",
    "max_pierce": "穿透次数",
    "aoe_radius": "AOE半径",
    "buff_on_hit": "命中Buff",
    "buff_chance": "Buff概率",
    "buff_on_self": "自身Buff",
    "duration": "持续时间",
    "interval": "跳数间隔",
    "tick_damage": "每跳伤害",
    "slow_ratio": "减速比例",
    "max_stacks": "最大层数",
    "effect_scene": "特效场景",
    "scene_path": "场景路径",
    "spawn_x": "出生点X",
    "spawn_y": "出生点Y",
    "bgm": "背景音乐",
    "asset": "资源目录",
    "character_config": "配置路径",
    "attack_range": "攻击范围",
    "detect_range": "检测范围",
    "patrol_range": "巡逻范围",
    "skills": "技能列表",
    "skill_weights": "技能权重",
    "drop_items": "掉落物品",
    "exp": "经验值",
}


def flatten_dict(d, prefix=""):
    """展平嵌套字典"""
    result = {}
    for key, value in d.items():
        full_key = f"{prefix}{key}" if not prefix else f"{prefix}_{key}"
        if isinstance(value, dict):
            result.update(flatten_dict(value, full_key))
        elif isinstance(value, list):
            result[full_key] = "; ".join(str(item) for item in value)
        else:
            result[full_key] = value
    return result


def get_header_display(key):
    """获取表头显示名（优先中文）"""
    return KEY_TO_CN.get(key, key)


def convert_json_to_excel(json_name, force=False):
    """转换单个 JSON 文件为 Excel"""
    json_path = os.path.join(JSON_DIR, f"{json_name}.json")
    if not os.path.exists(json_path):
        print(f"  跳过: {json_path} 不存在")
        return False

    excel_path = os.path.join(EXCEL_DIR, f"{json_name}.xlsx")
    if os.path.exists(excel_path) and not force:
        print(f"  跳过: {excel_path} 已存在（加 --force 覆盖）")
        return False

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        print(f"  跳过: {json_name}.json 不是字典格式")
        return False

    # 收集所有 key
    all_keys = []
    flat_rows = {}
    for entry_id, entry in data.items():
        if isinstance(entry, dict):
            flat = flatten_dict(entry)
            flat_rows[entry_id] = flat
            for key in flat:
                if key not in all_keys:
                    all_keys.append(key)
        else:
            flat_rows[entry_id] = {"value": entry}
            if "value" not in all_keys:
                all_keys.append("value")

    if not all_keys:
        print(f"  跳过: {json_name}.json 无可展开的字段")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = json_name.capitalize()

    # 写表头
    headers = ["id"] + [k for k in all_keys if k != "id"]
    for col, key in enumerate(headers, 1):
        display = get_header_display(key)
        cell = ws.cell(row=1, column=col, value=display)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = max(10, len(str(display)) * 2 + 4)

    # 写数据
    for row_idx, (entry_id, flat) in enumerate(flat_rows.items(), 2):
        ws.cell(row=row_idx, column=1, value=entry_id).border = THIN_BORDER
        for col_idx, key in enumerate(headers[1:], 2):
            val = flat.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    if ws.dimensions:
        ws.auto_filter.ref = ws.dimensions

    os.makedirs(EXCEL_DIR, exist_ok=True)
    wb.save(excel_path)
    print(f"  已生成: {excel_path}")
    return True


def main():
    force = "--force" in sys.argv
    args = [a for a in sys.argv[1:] if a != "--force"]

    if args:
        names = args
    else:
        names = []
        for f in os.listdir(JSON_DIR):
            if f.endswith(".json") and not f.startswith(".") and not f.startswith("~"):
                names.append(f[:-5])

    if not names:
        print(f"没有找到 JSON 文件: {JSON_DIR}")
        return

    print(f"JSON → Excel 转换")
    print(f"  JSON  目录: {JSON_DIR}")
    print(f"  Excel 目录: {EXCEL_DIR}")
    print()

    success = 0
    for name in names:
        print(f"转换: {name}.json")
        if convert_json_to_excel(name, force):
            success += 1

    print(f"\n完成: {success}/{len(names)} 个文件已转换")


if __name__ == "__main__":
    main()
